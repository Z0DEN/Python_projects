from PIL import Image
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def crop(file_name):
    filename = file_name
    with Image.open(filename) as img:
        img.load()
    print(img.size)
    cropped_img = img.crop((80, 100, 180, 200))
    print(cropped_img.size)
    cropped_img.save("cropped_image.jpg")

#crop('/home/blesk/GitHub/Python_projects/Excel/test.jpg')

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet")

im = Image.open("/home/blesk/GitHub/Python_projects/Excel/test.jpg")
pixels = im.load() # список с пикселями
x, y = im.size # ширина (x) и высота (y) изображения

process = '#'
count = 0
deg = 0.05
degree = 5
space = '------------------------'
k = 1
for i in range(x):
    for j in range(y):
        r, g, b = pixels[i, j]
        rgb = f"{r}, {g}, {b}"
        hex = (''.join([f"{int(i):02x}" for i in rgb.split(',')]))
        ws.cell(row=j+1, column=i+1).fill = PatternFill(start_color=f'{hex}', end_color=f'{hex}', fill_type="solid",)
        if ((i*j)/(x*y))>deg:
            print('process:','[',process,space[k:19],'] ---',degree,'%')
            deg += 0.05
            degree += 5
            process += '#'
            k += 1
wb.save('img-to-xlsx.xlsx')
pro = '####################'
print('process:','[',pro,']','--- 100 %')