from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = 'RowColDimension'
ws.column_dimensions['B'].width = 30
ws.column_dimensions['D'].width = 50
ws.row_dimensions[5].height = 50
ws['B5'] = 'ширина = 30, высота = 50'
ws['D5'] = 'ширина = 50, высота = 50'
from openpyxl.styles import Alignment
for cell in ws[5]:
    if cell.value:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        wb.save('width-height.xlsx')