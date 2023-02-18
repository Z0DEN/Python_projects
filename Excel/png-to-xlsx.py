from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet")
#wb = opx.load_workbook('pattern.xlsx')
cell = wb.active.cell(column=1, row=1)
ws['A4'] = 5

wb.save('pattern.xlsx')