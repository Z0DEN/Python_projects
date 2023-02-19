import openpyxl
from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np

alphabet = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet")

for j in range(1, 21):
    for i in range(1, 21):
            ws.cell(row=j, column=i).fill = PatternFill(start_color='FF91D2FF', end_color='FF91D2FF', fill_type="solid")
wb.save('pixel_art.xlsx')

input("Press Enter to continue...")

im = Image.new('RGB', (20, 20), (255, 255, 255))
draw = ImageDraw.Draw(im)

workbook = openpyxl.load_workbook('/home/blesk/GitHub/Python_projects/pixel_art.xlsx')
sheets_list = workbook.sheetnames
sheet_active = workbook[sheets_list[0]]
x = 0
y = 0
for cell_letter in alphabet:
    x += 1
    y = 0
    for cell_number in range(1, 21):
        cell_fill = sheet_active[f'{cell_letter}{cell_number}'].fill.start_color.index #Получаем цвет ячейки
        cell_hex ='#' + cell_fill[2:8]
        draw.rectangle((x-1, y, x-1, y), fill=f'{cell_hex}')
        y += 1
im.show()